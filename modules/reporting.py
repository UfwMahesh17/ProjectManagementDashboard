"""
reporting.py — Professional Excel report generation with openpyxl.
"""

from __future__ import annotations
import io
from datetime import date
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from modules.core import DepartmentResult, BASE_MARKS

DARK_NAVY  = "0D1B2A"
STEEL_BLUE = "1B4F72"
TEAL       = "1ABC9C"
LIGHT_ROW  = "EBF5FB"
WHITE      = "FFFFFF"
RED        = "E74C3C"
HEADER_FG  = "FFFFFF"

thin   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _cell_color(marks):
    if marks is None: return WHITE
    if marks >= 80:   return "D5F5E3"
    if marks >= 50:   return "FDEBD0"
    return "FADBD8"


def _hdr(ws, row, cols, bg=STEEL_BLUE):
    for col in cols:
        c = ws[f"{col}{row}"]
        c.font      = Font(bold=True, color=HEADER_FG, name="Calibri", size=10)
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = BORDER


def _data(ws, row, cols, bg=WHITE):
    for col in cols:
        c = ws[f"{col}{row}"]
        c.font      = Font(name="Calibri", size=10)
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = BORDER


def generate_report(
    all_projects: list[dict],   # list of {project_code, project_start, dept_results}
) -> bytes:
    wb = Workbook()

    # Sheet 1 — All Projects Overview
    ws_proj = wb.active
    ws_proj.title = "Projects Overview"
    _build_projects_overview(ws_proj, all_projects)

    # One sheet per project: Part Detail + Delay Log
    for proj in all_projects:
        code    = proj["project_code"]
        results = proj["dept_results"]

        ws_d = wb.create_sheet(f"{code} – Parts")
        _build_detail(ws_d, code, results)

        ws_l = wb.create_sheet(f"{code} – Delays")
        _build_delay_log(ws_l, code, results)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _build_projects_overview(ws, all_projects):
    # Title
    last_col = "L"
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "ADWIK INTELLIMECH — ALL PROJECTS OVERVIEW"
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill      = PatternFill("solid", start_color=DARK_NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = [
        "Project Code", "Start Date", "Departments",
        "Total Duration\n(days)", "Parts\nTracked", "Parts\nComplete",
        "Total Score", "Efficiency %",
        "Avg Dept Delay\n(days)", "Total Delay\n(days)", "Status",
    ]
    cols = list("ABCDEFGHIJK")
    ws.row_dimensions[3].height = 36
    for col, h in zip(cols, headers):
        ws[f"{col}3"] = h
    _hdr(ws, 3, cols)

    for r_idx, proj in enumerate(all_projects, start=4):
        results  = proj["dept_results"]
        code     = proj["project_code"]
        p_start  = proj["project_start"]
        bg       = LIGHT_ROW if r_idx % 2 == 0 else WHITE

        dept_count   = len(results)
        total_dur    = sum(dr.duration for dr in results)
        total_parts  = sum(len(dr.parts) for dr in results)
        done_parts   = sum(1 for dr in results for p in dr.parts if p.actual_finish)
        marks_list   = [dr.avg_marks for dr in results]
        total_score  = sum(marks_list)  # Sum all department scores
        eff_pct      = f"{(sum(marks_list) / (len(marks_list) * 100) * 100):.1f}%" if marks_list else "0%"
        
        # Calculate Average Department Delay
        dept_delays = [dr.actual_delay_out for dr in results]
        avg_dept_delay = sum(dept_delays) / len(dept_delays) if dept_delays else 0
        
        total_delay  = sum(dept_delays)
        status       = "✅ On Track" if total_delay == 0 else f"⚠️ {total_delay}d delay"

        row_data = [
            code,
            p_start.strftime("%d %b %Y") if isinstance(p_start, date) else str(p_start),
            dept_count, total_dur, total_parts, done_parts,
            total_score, eff_pct, round(avg_dept_delay, 1), total_delay, status,
        ]
        for col, val in zip(cols, row_data):
            ws[f"{col}{r_idx}"] = val
        _data(ws, r_idx, cols, bg)
        ws[f"G{r_idx}"].fill = PatternFill("solid", start_color=_cell_color(total_score / len(results) if results else 100))

    widths = [16, 14, 13, 14, 10, 12, 11, 12, 15, 13, 16]
    for col, w in zip(cols, widths):
        ws.column_dimensions[col].width = w

    # Bar chart of efficiency per project
    last_row = 3 + len(all_projects)
    chart = BarChart()
    chart.type    = "col"
    chart.title   = "Project Efficiency %"
    chart.y_axis.title = "Efficiency %"
    chart.style   = 10
    chart.width   = 20
    chart.height  = 12
    data_ref = Reference(ws, min_col=8, min_row=3, max_row=last_row)
    cats_ref = Reference(ws, min_col=1, min_row=4, max_row=last_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, f"A{last_row + 3}")


def _build_detail(ws, project_code, dept_results):
    ws.merge_cells("A1:L1")
    ws["A1"] = f"PART-LEVEL DETAIL — {project_code}"
    ws["A1"].font      = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    ws["A1"].fill      = PatternFill("solid", start_color=DARK_NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = [
        "Department", "Part Name", "MC", "Description", "PIC",
        "Planned Start", "Planned End", "Plan Dur (days)",
        "Actual Start", "Actual Finish", "Act Dur (days)",
        "Delay (days)", "Penalty Marks",
        "Rating"
    ]
    cols = list("ABCDEFGHIJKLMN")
    ws.row_dimensions[3].height = 36
    for col, h in zip(cols, headers):
        ws[f"{col}3"] = h
    _hdr(ws, 3, cols)

    row_num = 4
    for dr in sorted(dept_results, key=lambda d: d.order):
        for part in dr.parts:
            bg = LIGHT_ROW if row_num % 2 == 0 else WHITE
            
            # Plan Duration
            p_dur = (part.planned_end - part.planned_start).days + 1 if part.planned_start and part.planned_end else 0
            # Actual Duration
            a_dur = (part.actual_finish - part.actual_start).days + 1 if hasattr(part, "actual_start") and part.actual_start and part.actual_finish else 0

            # Use part.planned_end for delay calculation consistency
            deadline = part.planned_end if part.planned_end else dr.shifted_end
            p_delay = (part.actual_finish - deadline).days if part.actual_finish else 0
            p_delay = max(0, p_delay)
            
            # Check for external categories correctly
            is_external = part.delay_category in {"Client Approval Lag", "Client Change Request"}
            penalty = p_delay * 0.5 if not is_external else 0

            row_data = [
                dr.name, 
                part.name,
                part.mc,
                part.description,
                getattr(part, "pic", ""),
                part.planned_start.strftime("%d %b %Y") if part.planned_start else "—",
                deadline.strftime("%d %b %Y")           if deadline           else "—",
                p_dur if p_dur > 0 else "—",
                part.actual_start.strftime("%d %b %Y") if getattr(part, "actual_start", None) else "—",
                part.actual_finish.strftime("%d %b %Y") if part.actual_finish else "Pending",
                a_dur if a_dur > 0 else "—",
                p_delay if part.actual_finish else "—",
                penalty if part.actual_finish else "—",
                "✅ Good" if p_delay == 0 else "⚠️ Delay" if p_delay < 5 else "❌ Severe"
            ]
            for col, val in zip(cols, row_data):
                ws[f"{col}{row_num}"] = val
            _data(ws, row_num, cols, bg)
            row_num += 1

            # If Design dept, render explicit delay events under each part
            if dr.name.lower() == "design":
                for ev in getattr(part, "delay_events", []):
                    ev_bg = WHITE
                    ev_row = [
                        "",  # Department (left blank for event rows)
                        f"  {ev.type}",
                        getattr(ev, "pic", ""),
                        ev.start.strftime("%d %b %Y") if getattr(ev, "start", None) else "—",
                        ev.end.strftime("%d %b %Y")   if getattr(ev, "end", None)   else "—",
                        ev.days if ev.days > 0 else "—",
                        "—", "—", "—",
                        ev.days if ev.days > 0 else "—",
                        "—",
                        "Event",
                    ]
                    for col, val in zip(cols, ev_row):
                        ws[f"{col}{row_num}"] = val
                    _data(ws, row_num, cols, ev_bg)
                    row_num += 1

    widths = [16, 16, 10, 18, 12, 14, 14, 12, 14, 14, 12, 12, 14, 12]
    for col, w in zip(cols, widths):
        ws.column_dimensions[col].width = w


def _build_delay_log(ws, project_code, dept_results):
    ws.merge_cells("A1:G1")
    ws["A1"] = f"DELAY LOG — {project_code}"
    ws["A1"].font      = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    ws["A1"].fill      = PatternFill("solid", start_color=DARK_NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["Department", "Part Name", "MC", "Description", "PIC", "Delay Days",
               "Category", "Type", "Penalty Applied", "Reason / Notes"]
    cols = list("ABCDEFGHIJ")
    ws.row_dimensions[3].height = 36
    for col, h in zip(cols, headers):
        ws[f"{col}3"] = h
    _hdr(ws, 3, cols, bg=RED)

    row_num = 4
    for dr in sorted(dept_results, key=lambda d: d.order):
        for part in dr.parts:
            if part.actual_finish and part.delay_days > 0:
                bg      = "FEF9E7" if part.is_external else "FADBD8"
                penalty = "None (External)" if part.is_external else f"−{part.delay_days * 0.5} marks"
                row_data = [
                    dr.name, part.name, part.mc, part.description, getattr(part, "pic", ""), part.delay_days,
                    part.delay_category or "—",
                    "External" if part.is_external else "Internal",
                    penalty,
                    part.delay_reason or "—",
                ]
                for col, val in zip(cols, row_data):
                    ws[f"{col}{row_num}"] = val
                _data(ws, row_num, cols, bg)
                row_num += 1

            # Also list explicit delay events (e.g., Rework, Missed out drawing)
            for ev in getattr(part, "delay_events", []):
                if ev.days > 0:
                    ev_bg = "FFF8E1"
                    penalty = "—"
                    row_data = [
                        dr.name,
                        f"{part.name} — {ev.type}",
                        part.mc,
                        part.description,
                        getattr(ev, "pic", ""),
                        ev.days,
                        part.delay_category or "—",
                        ev.type,
                        penalty,
                        part.delay_reason or ev.notes or "—",
                    ]
                    for col, val in zip(cols, row_data):
                        ws[f"{col}{row_num}"] = val
                    _data(ws, row_num, cols, ev_bg)
                    row_num += 1

    if row_num == 4:
        ws.merge_cells("A4:J4")
        ws["A4"] = "✓  No delays recorded."
        ws["A4"].font      = Font(bold=True, color="27AE60", name="Calibri")
        ws["A4"].alignment = Alignment(horizontal="center")

    widths = [14, 16, 10, 16, 12, 10, 18, 12, 16, 30]
    for col, w in zip(cols, widths):
        ws.column_dimensions[col].width = w
